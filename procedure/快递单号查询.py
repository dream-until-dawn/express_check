from requests_html import HTMLSession
import openpyxl

import random
import os
import ast
import time
import re

file_path = os.path.realpath(__file__)

head_dic={#同样请求头
    'user-agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36',
}

excel_data = []

def chaxun(danghao_list):
    max_len = len(danghao_list)
    print('共有{}个单号，开始查询!'.format(max_len))
    session = HTMLSession()
    baidu_url = 'https://www.baidu.com/s?wd=318299071191910&rsv_spt=1&rsv_iqid=0xb3e59d420005676a&issp=1&f=8&rsv_bp=1&rsv_idx=2&ie=utf-8&tn=02003390_19_hao_pg&rsv_enter=0&rsv_dl=tb&rsv_sug3=2&rsv_sug1=2&rsv_sug7=101&rsv_n=2&rsv_btype=i&inputT=1338&rsv_sug4=1449'
    resp=session.get(baidu_url,headers=head_dic)
    #print(resp)
    resp.elapsed='utf-8'
    str_tokenV2 = re.findall(r'&tokenV2=(.*?)",',resp.text)
    print('第1次查询,拿取tokenV2,值:{}'.format(str_tokenV2[0]))
    url='https://express.baidu.com/express/api/express'
    #-------------新建表格-------------
    # wb = openpyxl.Workbook()
    # sheet = wb.active
    row_count = 1
    #-------------新建表格-------------
    for danghao in danghao_list:
        p_data = {
            'query_from_srcid': '4001',
            'isBaiduBoxApp': '10002',
            'isWisePc': '10020',
            'tokenV2': str_tokenV2[0],#会随机
            'cb': 'jQuery110207418431069154023_1655217556274',#返回值有带-不重要不用随机
            'appid': '4001',
            'com': 'yunda',
            'nu': danghao,
            'vcode': '',
            'token': '',
            'qid': '82ba0ce000029efc',#会随机
            '_': '1655214370724',#会随机
        }
        if danghao[0:2] == 'YT':
            p_data['com']='yuantong'#圆通
        if danghao[0:2] == 'SF':
            p_data['com']='shunfeng'#顺丰
        if danghao[0:2] == '75':
            p_data['com']='zhongtong'#中通
        resp=session.get(url,headers=head_dic,params=p_data)
        print('查询返回值:{}'.format(resp))
        result_str = resp.text.encode('utf-8').decode('unicode_escape')
        start = result_str.find('(')
        user_dict = ast.literal_eval(result_str[start+1:len(result_str)-1])
        try:
            if user_dict['data']['info']['status'] == '1':
                #print(user_dict['data']['info']['context'][0])
                str_time = time.localtime(int(user_dict['data']['info']['context'][0]['time']))
                str_time = time.strftime("%Y-%m-%d %H:%M:%S",str_time)
                #print('时间:{}---信息:{}'.format(str_time,user_dict['data']['info']['context'][0]['desc']))
                #-------------写入表格-------------
                # sheet.cell(row_count,1,danghao)
                # sheet.cell(row_count,2,str_time)
                # sheet.cell(row_count,3,user_dict['data']['info']['context'][0]['desc'])
                excel_data.append({'danhao':danghao,'type':p_data['com'],'time':str_time,'message':user_dict['data']['info']['context'][0]['desc']})
                row_count += 1
                print('单号{}查询成功!进度({}/{})'.format(danghao,row_count-1,max_len))
                #-------------写入表格-------------
            else:
                print('错误信息:{}'.format(user_dict))
                row_count += 1
        except:
            print('单号{}查询失败!进度({}/{})'.format(danghao,row_count-1,max_len))
            excel_data.append({'danhao':danghao,'type':p_data['com'],'time':'','message':''})
            row_count += 1

    resp.close()#关闭
    session.close()
    #-------------保存表格-------------
    # str_time = time.localtime(int(time.time()))
    # str_time = time.strftime("%Y年%m月%d日%H时%M分%S秒",str_time)
    # wb.save(file_path+'/../../'+str_time+'.xlsx')
    #-------------保存表格-------------
    return

def excel_new():
    wb = openpyxl.Workbook()
    sheet = wb.active
    row_count =1
    for user in excel_data:
        sheet.cell(row_count,1,user['danhao'])
        sheet.cell(row_count,2,user['type'])
        sheet.cell(row_count,3,user['time'])
        sheet.cell(row_count,4,user['message'])
        row_count += 1
    str_time = time.localtime(int(time.time()))
    str_time = time.strftime("查询结果-%Y年%m月%d日%H时%M分%S秒",str_time)
    wb.save(file_path+'/../../'+str_time+'.xlsx')
    return
#---------------------------------------------------------执行-----------------------------------------------------------
danghao_list = []
file_obj = open(file_path+'/../../单号放这里.txt')
all_lines = file_obj.readlines()
for line in all_lines:
    str_line = line.replace('\n','')
    danghao_list.append(str_line)
file_obj.close()

max_page = int(len(danghao_list)/50)
print(max_page)
print(len(danghao_list))
for page in range(0,max_page+1):
    try:
        print(len(danghao_list[page*50:page*50+50]))
        chaxun(danghao_list=danghao_list[page*50:page*50+50])
        #print(danghao_list[page*50:page*50+50])
    except:
        print(len(danghao_list[page*50:page*50+len(danghao_list)%50]))
        chaxun(danghao_list=danghao_list[page*50:page*50+len(danghao_list)%50])
        #print(danghao_list[page*50:page*50+len(danghao_list)%50])

#chaxun(danghao_list=danghao_list[0:50])
#chaxun(danghao_list=danghao_list[50:60])
#chaxun(danghao_list=danghao_list[100:120])
excel_new()